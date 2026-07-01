import calendar
import json
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


MONTHS = ["FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC", "ENE"]
MONTH_NUM_TO_KEY = {
    2: "FEB",
    3: "MAR",
    4: "ABR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AGO",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DIC",
    1: "ENE",
}


def as_str(value):
    return "" if value is None else str(value).strip()


def as_float(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = as_str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def normalize_header(value):
    text = as_str(value).upper()
    replacements = {
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "Ü": "U",
        "Ñ": "N",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def normalize_text_key(value):
    text = unicodedata.normalize("NFKD", as_str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def build_normalized_headers(headers):
    return [normalize_header(h) for h in headers]


def find_col(headers_norm, patterns, fallback=None):
    for token_group in patterns:
        for idx, header in enumerate(headers_norm):
            if all(token in header for token in token_group):
                return idx
    return fallback


def month_col_map(headers_norm):
    mapping = {}
    for month in MONTHS:
        for idx, header in enumerate(headers_norm):
            if month in header and month not in mapping:
                mapping[month] = idx
                break
    return mapping


def month_num_to_key(month_num):
    return MONTH_NUM_TO_KEY.get(month_num)


def build_period_label(ytd_months):
    if not ytd_months:
        return "Sin datos reales"
    if len(ytd_months) == 1:
        return ytd_months[0].title()
    return f"{ytd_months[0].title()} - {ytd_months[-1].title()}"


def resolve_data_dir(base_dir):
    for candidate in ("datos", "Datos"):
        path = base_dir / candidate
        if path.exists() and path.is_dir():
            return path
    raise FileNotFoundError("No se encontro la carpeta de datos (datos o Datos).")


def easter_sunday(year):
    # Computus (Gregorian calendar).
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def default_holidays_for_year(year):
    # Festivos nacionales habituales en Espana (aprox) + Viernes Santo.
    fixed = {
        date(year, 1, 1),
        date(year, 1, 6),
        date(year, 5, 1),
        date(year, 8, 15),
        date(year, 10, 12),
        date(year, 11, 1),
        date(year, 12, 6),
        date(year, 12, 8),
        date(year, 12, 25),
    }
    easter = easter_sunday(year)
    fixed.add(easter - timedelta(days=2))  # Viernes Santo.
    return fixed


def load_holidays(data_dir, year):
    # Ajuste temporal: previsión solo con exclusión de sábados y domingos.
    return set()


def is_working_day(check_date, holidays):
    return check_date.weekday() < 5 and check_date not in holidays


def count_working_days(start_date, end_date, holidays):
    if start_date > end_date:
        return 0
    count = 0
    cursor = start_date
    while cursor <= end_date:
        if is_working_day(cursor, holidays):
            count += 1
        cursor += timedelta(days=1)
    return count


def find_file_case_insensitive(folder, target_name):
    target = target_name.lower()
    for item in folder.iterdir():
        if item.is_file() and item.name.lower() == target:
            return item
    raise FileNotFoundError(f"No se encontro {target_name} en {folder}")


def find_optional_file_case_insensitive(folder, target_name):
    target = target_name.lower()
    for item in folder.iterdir():
        if item.is_file() and item.name.lower() == target:
            return item
    return None


def find_budget_file(data_dir):
    preferred_names = ("PRESUPUESTOS.xlsx", "PRESUPUESTOS VENTAS.xlsx")

    for name in preferred_names:
        try:
            return find_file_case_insensitive(data_dir, name)
        except FileNotFoundError:
            pass

    raise FileNotFoundError(
        "No se encontro archivo de presupuestos en la carpeta de datos "
        "(nombres esperados: PRESUPUESTOS.xlsx o PRESUPUESTOS VENTAS.xlsx)."
    )


def find_franchises_file(data_dir):
    preferred_names = ("FRANQUICIAS.xlsx", "FRANQUICIAS.xlsm")
    for name in preferred_names:
        found = find_optional_file_case_insensitive(data_dir, name)
        if found:
            return found
    return None


def init_agent(agents_data, agent_id, name):
    if agent_id not in agents_data:
        agents_data[agent_id] = {
            "id": agent_id,
            "name": as_str(name) or f"Agente {agent_id}",
            "sales_2025_monthly": {m: 0.0 for m in MONTHS},
            "budget_2026_monthly": {m: 0.0 for m in MONTHS},
            "sales_2026_monthly": {m: 0.0 for m in MONTHS},
            "sales_2025_daily": {},
            "profit_2026_monthly": {m: 0.0 for m in MONTHS},
            "sales_2026_daily": {},
            "profit_2026_daily": {},
            "total_sales_2025": 0.0,
            "total_budget_2026": 0.0,
            "sales_2025_ytd": 0.0,
            "budget_2026_ytd": 0.0,
            "sales_2026_ytd": 0.0,
            "profit_2026_ytd": 0.0,
            "deviation_pct": 0.0,
            "growth_pct": 0.0,
            "margin_pct": 0.0,
            "forecast_sales_month_end": 0.0,
            "forecast_profit_month_end": 0.0,
            "forecast_margin_pct_month_end": 0.0,
            "expected_compliance_pct_month_end": 0.0,
            "required_daily_sales_to_budget": 0.0,
            "clients": {},
        }
    return agents_data[agent_id]


def init_client(agent, client_id, client_name):
    client_key = str(client_id) if client_id is not None else f"NAME::{as_str(client_name).upper()}"
    if client_key not in agent["clients"]:
        agent["clients"][client_key] = {
            "id": client_id if client_id is not None else "N/A",
            "name": as_str(client_name) or "Cliente sin nombre",
            "sales_2025": 0.0,
            "budget_2026": 0.0,
            "sales_2026": 0.0,
            "profit_2026": 0.0,
            "sales_2025_monthly": {m: 0.0 for m in MONTHS},
            "budget_2026_monthly": {m: 0.0 for m in MONTHS},
            "sales_2026_monthly": {m: 0.0 for m in MONTHS},
            "profit_2026_monthly": {m: 0.0 for m in MONTHS},
            "sales_2025_ytd": 0.0,
            "budget_2026_ytd": 0.0,
            "sales_2026_ytd": 0.0,
            "profit_2026_ytd": 0.0,
            "deviation_pct": 0.0,
            "growth_pct": 0.0,
            "margin_pct": 0.0,
        }
    return agent["clients"][client_key]


def add_daily_value(target_map, key_date, value):
    key = key_date.isoformat()
    target_map[key] = target_map.get(key, 0.0) + value


def compute_month_forecast(sales_daily, profit_daily, as_of_date, month_budget, holidays=None):
    holidays = holidays or set()
    if not as_of_date:
        return {
            "forecast_sales_month_end": 0.0,
            "forecast_profit_month_end": 0.0,
            "forecast_margin_pct_month_end": 0.0,
            "expected_compliance_pct_month_end": 0.0,
            "required_daily_sales_to_budget": 0.0,
            "current_month_sales": 0.0,
            "current_month_profit": 0.0,
            "days_elapsed": 0,
            "days_in_month": 0,
            "days_remaining": 0,
        }

    year = as_of_date.year
    month = as_of_date.month
    days_in_month = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, days_in_month)

    # Prevision basada en dias laborables (sin sabados, domingos ni festivos).
    days_elapsed = count_working_days(month_start, as_of_date, holidays)
    days_remaining = count_working_days(as_of_date + timedelta(days=1), month_end, holidays)

    month_prefix = f"{year:04d}-{month:02d}-"
    current_month_sales = sum(value for day, value in sales_daily.items() if day.startswith(month_prefix))
    current_month_profit = sum(value for day, value in profit_daily.items() if day.startswith(month_prefix))

    avg_daily_sales = current_month_sales / days_elapsed if days_elapsed > 0 else 0.0
    avg_daily_profit = current_month_profit / days_elapsed if days_elapsed > 0 else 0.0

    recent_working_days = []
    cursor = as_of_date
    while len(recent_working_days) < 7 and cursor.month == month:
        if is_working_day(cursor, holidays):
            recent_working_days.append(cursor)
        cursor -= timedelta(days=1)

    recent_sales = sum(sales_daily.get(d.isoformat(), 0.0) for d in recent_working_days)
    recent_profit = sum(profit_daily.get(d.isoformat(), 0.0) for d in recent_working_days)
    divisor_recent = len(recent_working_days)

    avg_recent_sales = recent_sales / divisor_recent if divisor_recent > 0 else 0.0
    avg_recent_profit = recent_profit / divisor_recent if divisor_recent > 0 else 0.0

    weighted_sales = 0.7 * avg_recent_sales + 0.3 * avg_daily_sales
    weighted_profit = 0.7 * avg_recent_profit + 0.3 * avg_daily_profit

    forecast_sales = current_month_sales + (weighted_sales * days_remaining)
    forecast_profit = current_month_profit + (weighted_profit * days_remaining)

    forecast_margin_pct = (forecast_profit / forecast_sales) * 100 if forecast_sales > 0 else 0.0
    expected_compliance_pct = (forecast_sales / month_budget) * 100 if month_budget > 0 else 0.0
    required_daily_sales = (
        max(0.0, month_budget - current_month_sales) / days_remaining if days_remaining > 0 else 0.0
    )

    return {
        "forecast_sales_month_end": forecast_sales,
        "forecast_profit_month_end": forecast_profit,
        "forecast_margin_pct_month_end": forecast_margin_pct,
        "expected_compliance_pct_month_end": expected_compliance_pct,
        "required_daily_sales_to_budget": required_daily_sales,
        "current_month_sales": current_month_sales,
        "current_month_profit": current_month_profit,
        "days_elapsed": days_elapsed,
        "days_in_month": days_in_month,
        "days_remaining": days_remaining,
    }


def compute_derived_metrics(agents_data, ytd_months, as_of_date, current_month_key, holidays=None):
    for agent in agents_data.values():
        agent["sales_2025_ytd"] = sum(agent["sales_2025_monthly"][m] for m in ytd_months)
        agent["budget_2026_ytd"] = sum(agent["budget_2026_monthly"][m] for m in ytd_months)
        agent["sales_2026_ytd"] = sum(agent["sales_2026_monthly"][m] for m in ytd_months)
        agent["profit_2026_ytd"] = sum(agent["profit_2026_monthly"][m] for m in ytd_months)

        agent["total_sales_2025"] = sum(agent["sales_2025_monthly"][m] for m in MONTHS)
        agent["total_budget_2026"] = sum(agent["budget_2026_monthly"][m] for m in MONTHS)

        if agent["budget_2026_ytd"] > 0:
            agent["deviation_pct"] = ((agent["sales_2026_ytd"] - agent["budget_2026_ytd"]) / agent["budget_2026_ytd"]) * 100
        if agent["sales_2025_ytd"] > 0:
            agent["growth_pct"] = ((agent["sales_2026_ytd"] - agent["sales_2025_ytd"]) / agent["sales_2025_ytd"]) * 100
        if agent["sales_2026_ytd"] > 0:
            agent["margin_pct"] = (agent["profit_2026_ytd"] / agent["sales_2026_ytd"]) * 100

        forecast_metrics = compute_month_forecast(
            agent["sales_2026_daily"],
            agent["profit_2026_daily"],
            as_of_date,
            agent["budget_2026_monthly"].get(current_month_key, 0.0),
            holidays,
        )
        agent.update({
            "forecast_sales_month_end": forecast_metrics["forecast_sales_month_end"],
            "forecast_profit_month_end": forecast_metrics["forecast_profit_month_end"],
            "forecast_margin_pct_month_end": forecast_metrics["forecast_margin_pct_month_end"],
            "expected_compliance_pct_month_end": forecast_metrics["expected_compliance_pct_month_end"],
            "required_daily_sales_to_budget": forecast_metrics["required_daily_sales_to_budget"],
        })

        clients = list(agent["clients"].values())
        for client in clients:
            client["sales_2025"] = sum(client["sales_2025_monthly"][m] for m in MONTHS)
            client["budget_2026"] = sum(client["budget_2026_monthly"][m] for m in MONTHS)
            client["sales_2026"] = sum(client["sales_2026_monthly"][m] for m in MONTHS)
            client["profit_2026"] = sum(client["profit_2026_monthly"][m] for m in MONTHS)

            client["sales_2025_ytd"] = sum(client["sales_2025_monthly"][m] for m in ytd_months)
            client["budget_2026_ytd"] = sum(client["budget_2026_monthly"][m] for m in ytd_months)
            client["sales_2026_ytd"] = sum(client["sales_2026_monthly"][m] for m in ytd_months)
            client["profit_2026_ytd"] = sum(client["profit_2026_monthly"][m] for m in ytd_months)

            if client["budget_2026"] > 0:
                client["deviation_pct"] = ((client["sales_2026"] - client["budget_2026"]) / client["budget_2026"]) * 100
            if client["sales_2025"] > 0:
                client["growth_pct"] = ((client["sales_2026"] - client["sales_2025"]) / client["sales_2025"]) * 100
            if client["sales_2026"] > 0:
                client["margin_pct"] = (client["profit_2026"] / client["sales_2026"]) * 100

        clients.sort(key=lambda c: (as_int(c["id"]) is None, as_int(c["id"]) or 10**12, c["name"]))
        agent["clients"] = clients


def process_budget_file(sheet, agents_data):
    headers = [cell.value for cell in sheet[1]]
    headers_norm = build_normalized_headers(headers)

    agent_id_col = find_col(headers_norm, [("AGENTE", "ID"), ("COD", "AGENTE"), ("AGENTE",)], fallback=0)
    agent_name_col = find_col(headers_norm, [("NOMBRE", "AGENTE"), ("AGENTE",)], fallback=1)
    client_id_col = find_col(headers_norm, [("CLIENTE", "ID"), ("CLI",), ("CLIENTE",)], fallback=2)
    client_name_col = find_col(headers_norm, [("NOMBRE", "CLIENTE"), ("RAZON", "SOCIAL")], fallback=3)
    type_col = find_col(headers_norm, [("TIPO",), ("CLASE",)], fallback=4)
    month_cols = month_col_map(headers_norm)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        agent_id = as_int(row[agent_id_col] if agent_id_col < len(row) else None)
        if agent_id is None:
            continue

        row_type = as_str(row[type_col] if type_col < len(row) else "")
        if not (row_type.startswith("PPTTO") or "PRESUP" in row_type):
            continue

        agent_name = row[agent_name_col] if agent_name_col < len(row) else f"Agente {agent_id}"
        client_id = as_int(row[client_id_col] if client_id_col < len(row) else None)
        client_name = row[client_name_col] if client_name_col < len(row) else f"Cliente {client_id or 'N/A'}"

        agent = init_agent(agents_data, agent_id, agent_name)
        client = init_client(agent, client_id, client_name)

        for month in MONTHS:
            month_col = month_cols.get(month)
            if month_col is None or month_col >= len(row):
                continue
            val = as_float(row[month_col])
            agent["budget_2026_monthly"][month] += val
            client["budget_2026_monthly"][month] += val


def process_daily_sales_file(sheet, year, agents_data):
    headers = [cell.value for cell in sheet[1]]
    usable_headers = headers[:-1] if len(headers) > 1 else headers
    headers_norm = build_normalized_headers(usable_headers)

    date_col = find_col(headers_norm, [("FECHA",), ("DATE",)], fallback=0)
    agent_id_col = find_col(headers_norm, [("AGENTE", "ID"), ("COD", "AGENTE"), ("AGENTE",)], fallback=1)
    agent_name_col = find_col(headers_norm, [("AGENTE", "VENTA"), ("NOMBRE", "AGENTE"), ("AGENTE",)], fallback=2)
    client_id_col = find_col(headers_norm, [("CLIENTE", "ID"), ("COD", "CLIENTE"), ("CLIENTE",)], fallback=3)
    client_name_col = find_col(headers_norm, [("NOMBRE", "CLIENTE"), ("RAZON", "SOCIAL")], fallback=4)
    importe_col = find_col(headers_norm, [("IMPORTE", "NETO"), ("IMPORTE", "TOTAL"), ("IMPORTE",)], fallback=8)
    beneficio_col = find_col(headers_norm, [("BENEFICIO", "TOTAL"), ("BENEFICIO",), ("GANANCIA",)], fallback=9)

    max_col = len(usable_headers)
    months_seen = set()
    max_date = None

    for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
        if not row:
            continue

        doc_date = as_date(row[date_col] if date_col < len(row) else None)
        if doc_date is None or doc_date.year != year:
            continue

        agent_id = as_int(row[agent_id_col] if agent_id_col < len(row) else None)
        if agent_id is None:
            continue

        agent_name = row[agent_name_col] if agent_name_col < len(row) else f"Agente {agent_id}"
        client_id = as_int(row[client_id_col] if client_id_col < len(row) else None)
        client_name = row[client_name_col] if client_name_col < len(row) else f"Cliente {client_id or 'N/A'}"
        importe = as_float(row[importe_col] if importe_col < len(row) else None)
        beneficio = as_float(row[beneficio_col] if beneficio_col < len(row) else None)

        month_key = month_num_to_key(doc_date.month)
        if month_key not in MONTHS:
            continue

        agent = init_agent(agents_data, agent_id, agent_name)
        client = init_client(agent, client_id, client_name)

        if as_str(agent_name) and len(as_str(agent_name)) >= len(agent["name"]):
            agent["name"] = as_str(agent_name)
        if as_str(client_name) and len(as_str(client_name)) >= len(client["name"]):
            client["name"] = as_str(client_name)

        if year == 2025:
            agent["sales_2025_monthly"][month_key] += importe
            add_daily_value(agent["sales_2025_daily"], doc_date, importe)
            client["sales_2025_monthly"][month_key] += importe
        elif year == 2026:
            agent["sales_2026_monthly"][month_key] += importe
            agent["profit_2026_monthly"][month_key] += beneficio
            add_daily_value(agent["sales_2026_daily"], doc_date, importe)
            add_daily_value(agent["profit_2026_daily"], doc_date, beneficio)

            client["sales_2026_monthly"][month_key] += importe
            client["profit_2026_monthly"][month_key] += beneficio

            if importe != 0 or beneficio != 0:
                months_seen.add(month_key)

            if max_date is None or doc_date > max_date:
                max_date = doc_date

    return months_seen, max_date


def build_global_daily_maps(agents_data):
    sales_daily = {}
    profit_daily = {}
    for agent in agents_data.values():
        for day, value in agent["sales_2026_daily"].items():
            sales_daily[day] = sales_daily.get(day, 0.0) + value
        for day, value in agent["profit_2026_daily"].items():
            profit_daily[day] = profit_daily.get(day, 0.0) + value
    return sales_daily, profit_daily


def build_global_sales_2025_daily_map(agents_data):
    sales_daily = {}
    for agent in agents_data.values():
        for day, value in agent["sales_2025_daily"].items():
            sales_daily[day] = sales_daily.get(day, 0.0) + value
    return sales_daily


def compute_same_date_sales_previous_year(sales_2025_daily, as_of_date):
    if not as_of_date:
        return 0.0

    prev_year = as_of_date.year - 1
    month = as_of_date.month
    max_day_prev_year = calendar.monthrange(prev_year, month)[1]
    cutoff_day = min(as_of_date.day, max_day_prev_year)

    prefix = f"{prev_year:04d}-{month:02d}-"
    return sum(
        value
        for day, value in sales_2025_daily.items()
        if day.startswith(prefix) and int(day[-2:]) <= cutoff_day
    )


def load_franchise_catalog(franchises_file):
    if not franchises_file:
        return {
            "enabled": False,
            "source_file": None,
            "entries_count": 0,
            "client_ids": set(),
            "client_names": set(),
        }

    wb = openpyxl.load_workbook(str(franchises_file), data_only=True)
    sheet = wb.active

    header_row_idx = 1
    header_values = [cell.value for cell in sheet[1]]
    for idx in range(1, 16):
        row_values = [cell.value for cell in sheet[idx]]
        norm = build_normalized_headers(row_values)
        if any("CLIENTE" in h for h in norm):
            header_row_idx = idx
            header_values = row_values
            break

    headers_norm = build_normalized_headers(header_values)
    client_id_col = find_col(headers_norm, [("CLIENTE", "ID"), ("COD", "CLIENTE"), ("ID",)], fallback=0)
    client_name_col = find_col(headers_norm, [("RAZON", "SOCIAL"), ("NOMBRE", "CLIENTE"), ("CLIENTE",)], fallback=1)
    active_col = find_col(headers_norm, [("ESTADO",), ("ACTIVO",), ("VIGENTE",)], fallback=None)

    client_ids = set()
    client_names = set()
    entries_count = 0

    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue

        state_txt = as_str(row[active_col] if active_col is not None and active_col < len(row) else "")
        state_norm = normalize_text_key(state_txt)
        if state_norm in {"BAJA", "INACTIVO", "INACTIVA", "NO"}:
            continue

        client_id = as_int(row[client_id_col] if client_id_col < len(row) else None)
        client_name = as_str(row[client_name_col] if client_name_col < len(row) else "")

        if client_id is None and not client_name:
            continue

        if client_id is not None:
            client_ids.add(client_id)
        if client_name:
            client_names.add(normalize_text_key(client_name))
        entries_count += 1

    return {
        "enabled": True,
        "source_file": franchises_file.name,
        "entries_count": entries_count,
        "client_ids": client_ids,
        "client_names": client_names,
    }


def compute_franchise_metrics(agents_data, global_totals, ytd_months, franchise_catalog):
    franchise_clients = {}
    franchise_ids = franchise_catalog["client_ids"]
    franchise_names = franchise_catalog["client_names"]

    def init_franchise_client(key, client):
        if key not in franchise_clients:
            franchise_clients[key] = {
                "id": client.get("id", "N/A"),
                "name": as_str(client.get("name", "Cliente sin nombre")),
                "sales_2025_monthly": {m: 0.0 for m in MONTHS},
                "sales_2026_monthly": {m: 0.0 for m in MONTHS},
                "budget_2026_monthly": {m: 0.0 for m in MONTHS},
            }
        return franchise_clients[key]

    for agent in agents_data.values():
        for client in agent.get("clients", []):
            client_id = as_int(client.get("id"))
            client_name_norm = normalize_text_key(client.get("name"))
            is_franchise = (client_id is not None and client_id in franchise_ids) or (client_name_norm in franchise_names)
            if not is_franchise:
                continue

            key = f"ID::{client_id}" if client_id is not None else f"NAME::{client_name_norm}"
            row = init_franchise_client(key, client)

            for month in MONTHS:
                row["sales_2025_monthly"][month] += as_float(client.get("sales_2025_monthly", {}).get(month, 0.0))
                row["sales_2026_monthly"][month] += as_float(client.get("sales_2026_monthly", {}).get(month, 0.0))
                row["budget_2026_monthly"][month] += as_float(client.get("budget_2026_monthly", {}).get(month, 0.0))

    monthly_rows = []
    franchise_2025_ytd = 0.0
    franchise_2026_ytd = 0.0

    for month in MONTHS:
        channel_2025 = as_float(global_totals["sales_2025_monthly"].get(month, 0.0))
        channel_2026 = as_float(global_totals["sales_2026_monthly"].get(month, 0.0))
        franchise_2025 = sum(c["sales_2025_monthly"][month] for c in franchise_clients.values())
        franchise_2026 = sum(c["sales_2026_monthly"][month] for c in franchise_clients.values())

        share_2025 = (franchise_2025 / channel_2025) * 100 if channel_2025 > 0 else 0.0
        share_2026 = (franchise_2026 / channel_2026) * 100 if channel_2026 > 0 else 0.0

        if month in ytd_months:
            franchise_2025_ytd += franchise_2025
            franchise_2026_ytd += franchise_2026

        monthly_rows.append({
            "month": month,
            "channel_2025": channel_2025,
            "franchise_2025": franchise_2025,
            "share_2025": share_2025,
            "channel_2026": channel_2026,
            "franchise_2026": franchise_2026,
            "share_2026": share_2026,
            "delta_share_pp": share_2026 - share_2025,
        })

    channel_2025_ytd = sum(as_float(global_totals["sales_2025_monthly"].get(m, 0.0)) for m in ytd_months)
    channel_2026_ytd = sum(as_float(global_totals["sales_2026_monthly"].get(m, 0.0)) for m in ytd_months)
    share_2025_ytd = (franchise_2025_ytd / channel_2025_ytd) * 100 if channel_2025_ytd > 0 else 0.0
    share_2026_ytd = (franchise_2026_ytd / channel_2026_ytd) * 100 if channel_2026_ytd > 0 else 0.0

    clients_rows = []
    franchise_budget_ytd = 0.0
    for client in franchise_clients.values():
        sales_2025_ytd = sum(client["sales_2025_monthly"][m] for m in ytd_months)
        sales_2026_ytd = sum(client["sales_2026_monthly"][m] for m in ytd_months)
        client_budget_ytd = sum(client["budget_2026_monthly"][m] for m in ytd_months)
        growth_pct = ((sales_2026_ytd - sales_2025_ytd) / sales_2025_ytd) * 100 if sales_2025_ytd > 0 else 0.0
        franchise_budget_ytd += client_budget_ytd
        clients_rows.append({
            "id": client["id"],
            "name": client["name"],
            "sales_2025_ytd": sales_2025_ytd,
            "sales_2026_ytd": sales_2026_ytd,
            "budget_2026_ytd": client_budget_ytd,
            "growth_pct": growth_pct,
            "sales_2025_monthly": client["sales_2025_monthly"],
            "sales_2026_monthly": client["sales_2026_monthly"],
            "budget_2026_monthly": client["budget_2026_monthly"],
        })

    clients_rows.sort(key=lambda c: c["sales_2026_ytd"], reverse=True)

    return {
        "enabled": franchise_catalog["enabled"],
        "source_file": franchise_catalog["source_file"],
        "catalog_entries": franchise_catalog["entries_count"],
        "matched_clients": len(clients_rows),
        "summary": {
            "franchise_2025_ytd": franchise_2025_ytd,
            "franchise_2026_ytd": franchise_2026_ytd,
            "channel_2025_ytd": channel_2025_ytd,
            "channel_2026_ytd": channel_2026_ytd,
            "share_2025_ytd": share_2025_ytd,
            "share_2026_ytd": share_2026_ytd,
            "delta_share_ytd_pp": share_2026_ytd - share_2025_ytd,
            "franchise_budget_ytd": franchise_budget_ytd,
        },
        "monthly": monthly_rows,
        "clients": clients_rows,
    }


def generate():
    base_dir = Path(__file__).resolve().parent
    data_dir = resolve_data_dir(base_dir)

    budget_file = find_budget_file(data_dir)
    sales_2025_file = find_file_case_insensitive(data_dir, "VENTAS_2025.xlsx")
    sales_2026_file = find_file_case_insensitive(data_dir, "VENTAS_2026.xlsx")
    franchises_file = find_franchises_file(data_dir)

    print("Starting data extraction...")
    print(f"Data directory: {data_dir}")
    print(f"Budget file: {budget_file}")

    agents_data = {}

    wb_budget = openpyxl.load_workbook(str(budget_file), data_only=True)
    sheet_budget = wb_budget.active
    process_budget_file(sheet_budget, agents_data)

    wb_2025 = openpyxl.load_workbook(str(sales_2025_file), data_only=True)
    sheet_2025 = wb_2025.active
    process_daily_sales_file(sheet_2025, 2025, agents_data)

    wb_2026 = openpyxl.load_workbook(str(sales_2026_file), data_only=True)
    sheet_2026 = wb_2026.active
    real_months_seen, as_of_date = process_daily_sales_file(sheet_2026, 2026, agents_data)

    ytd_months = [m for m in MONTHS if m in real_months_seen]
    current_month_key = month_num_to_key(as_of_date.month) if as_of_date else None

    holiday_year = as_of_date.year if as_of_date else datetime.now().year
    holidays = load_holidays(data_dir, holiday_year)

    compute_derived_metrics(agents_data, ytd_months, as_of_date, current_month_key, holidays)

    global_totals = {
        "sales_2025_monthly": {m: 0.0 for m in MONTHS},
        "budget_2026_monthly": {m: 0.0 for m in MONTHS},
        "sales_2026_monthly": {m: 0.0 for m in MONTHS},
        "profit_2026_monthly": {m: 0.0 for m in MONTHS},
        "total_sales_2025": 0.0,
        "total_budget_2026": 0.0,
        "sales_2025_ytd": 0.0,
        "budget_2026_ytd": 0.0,
        "sales_2026_ytd": 0.0,
        "profit_2026_ytd": 0.0,
    }

    for agent in agents_data.values():
        for month in MONTHS:
            global_totals["sales_2025_monthly"][month] += agent["sales_2025_monthly"][month]
            global_totals["budget_2026_monthly"][month] += agent["budget_2026_monthly"][month]
            global_totals["sales_2026_monthly"][month] += agent["sales_2026_monthly"][month]
            global_totals["profit_2026_monthly"][month] += agent["profit_2026_monthly"][month]

        global_totals["total_sales_2025"] += agent["total_sales_2025"]
        global_totals["total_budget_2026"] += agent["total_budget_2026"]
        global_totals["sales_2025_ytd"] += agent["sales_2025_ytd"]
        global_totals["budget_2026_ytd"] += agent["budget_2026_ytd"]
        global_totals["sales_2026_ytd"] += agent["sales_2026_ytd"]
        global_totals["profit_2026_ytd"] += agent["profit_2026_ytd"]

    global_sales_daily, global_profit_daily = build_global_daily_maps(agents_data)
    global_sales_2025_daily = build_global_sales_2025_daily_map(agents_data)
    global_forecast = compute_month_forecast(
        global_sales_daily,
        global_profit_daily,
        as_of_date,
        global_totals["budget_2026_monthly"].get(current_month_key, 0.0) if current_month_key else 0.0,
        holidays,
    )
    current_month_sales_2025_same_date = compute_same_date_sales_previous_year(global_sales_2025_daily, as_of_date)
    franchise_catalog = load_franchise_catalog(franchises_file)
    franchise_metrics = compute_franchise_metrics(agents_data, global_totals, ytd_months, franchise_catalog)

    current_ts = datetime.now()

    for agent in agents_data.values():
        agent.pop("sales_2025_daily", None)
        agent.pop("sales_2026_daily", None)
        agent.pop("profit_2026_daily", None)

    dashboard_data = {
        "last_updated": current_ts.strftime("%d/%m/%Y"),
        "as_of_date": as_of_date.isoformat() if as_of_date else None,
        "period_label": build_period_label(ytd_months),
        "ytd_months": ytd_months,
        "months": MONTHS,
        "fiscal_period": "01/02/2026 - 31/01/2027",
        "global_totals": global_totals,
        "forecast": {
            "current_month_key": current_month_key,
            "forecast_sales_month_end": global_forecast["forecast_sales_month_end"],
            "forecast_profit_month_end": global_forecast["forecast_profit_month_end"],
            "forecast_margin_pct_month_end": global_forecast["forecast_margin_pct_month_end"],
            "expected_compliance_pct_month_end": global_forecast["expected_compliance_pct_month_end"],
            "required_daily_sales_to_budget": global_forecast["required_daily_sales_to_budget"],
            "days_elapsed": global_forecast["days_elapsed"],
            "days_in_month": global_forecast["days_in_month"],
            "days_remaining": global_forecast["days_remaining"],
            "current_month_sales": global_forecast["current_month_sales"],
            "current_month_sales_2025_same_date": current_month_sales_2025_same_date,
            "current_month_profit": global_forecast["current_month_profit"],
        },
        "franchises": franchise_metrics,
        "agents": sorted(list(agents_data.values()), key=lambda agent: agent["id"]),
        "source_files": {
            "budget": budget_file.name,
            "sales_2025": sales_2025_file.name,
            "sales_2026": sales_2026_file.name,
            "franchises": franchises_file.name if franchises_file else None,
        },
    }

    output_path = base_dir / "data.js"
    with output_path.open("w", encoding="utf-8") as file_out:
        file_out.write("// Generado automaticamente por generate_data.py\n")
        file_out.write(f"const DASHBOARD_DATA = {json.dumps(dashboard_data, indent=2, ensure_ascii=False)};\n")

    print("Data extraction completed successfully! Written to data.js.")


if __name__ == "__main__":
    generate()
