import openpyxl


def as_str(value):
    return "" if value is None else str(value).strip()


def as_float(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_header(value):
    text = as_str(value).upper()
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def find_col(headers_norm, patterns, fallback=None):
    for token_group in patterns:
        for idx, header in enumerate(headers_norm):
            if all(token in header for token in token_group):
                return idx
    return fallback


def summarize_presupuestos(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb["CANAL_DI_MY_AGRUPADOS"]

    headers = [cell.value for cell in sheet[1]]
    headers_norm = [normalize_header(h) for h in headers]
    print("CANAL_DI_MY_AGRUPADOS Headers:", headers)

    agent_id_col = find_col(headers_norm, [("AGENTE", "ID"), ("COD", "AGENTE")], fallback=0)
    agent_name_col = find_col(headers_norm, [("AGENTE", "VENTA"), ("NOMBRE", "AGENTE"), ("AGENTE",)], fallback=1)
    client_id_col = find_col(headers_norm, [("CLIENTE", "ID"), ("COD", "CLIENTE"), ("CLIENTE",)], fallback=4)
    client_name_col = find_col(headers_norm, [("NOMBRE", "CLIENTE"), ("RAZON", "SOCIAL")], fallback=5)
    type_col = find_col(headers_norm, [("TIPO",), ("CLASE",)], fallback=4)

    agents = {}
    row_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        agent_id = as_int(row[agent_id_col] if agent_id_col < len(row) else None)
        if agent_id is None:
            continue

        row_count += 1
        agent_name = as_str(row[agent_name_col] if agent_name_col < len(row) else f"Agente {agent_id}")
        client_name = as_str(row[client_name_col] if client_name_col < len(row) else "")
        row_type = as_str(row[type_col] if type_col < len(row) else "")

        if agent_id not in agents:
            agents[agent_id] = {
                "name": agent_name,
                "total_ventas_2025": 0.0,
                "total_pptto_2026": 0.0,
                "clients": set(),
            }

        if client_name:
            agents[agent_id]["clients"].add(client_name)

        months_sum = 0.0
        for month in ["FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC", "ENE"]:
            month_col = next((idx for idx, h in enumerate(headers_norm) if month in h), None)
            if month_col is not None and month_col < len(row):
                months_sum += as_float(row[month_col])

        if row_type.startswith("VENTA"):
            agents[agent_id]["total_ventas_2025"] += months_sum
        elif row_type.startswith("PPTTO") or "PRESUP" in row_type:
            agents[agent_id]["total_pptto_2026"] += months_sum

    print(f"Total filas presupuestos: {row_count}")
    print(f"Agentes unicos en presupuestos: {len(agents)}")
    for aid, info in sorted(agents.items()):
        print(
            f"Agente {aid} ({info['name']}): "
            f"Ventas 2025={info['total_ventas_2025']:.2f}, "
            f"Ppto 2026={info['total_pptto_2026']:.2f}, "
            f"Clientes={len(info['clients'])}"
        )


def summarize_ventas_reales(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.active
    print(f"\nSheet name in {filename}: {sheet.title}")

    headers = [cell.value for cell in sheet[1]]
    usable_headers = headers[:-1] if len(headers) > 1 else headers
    headers_norm = [normalize_header(h) for h in usable_headers]

    print("Ventas 2026 Headers (usadas):", usable_headers)
    if len(headers) > len(usable_headers):
        print("Columna final ignorada:", headers[-1])

    agent_id_col = find_col(headers_norm, [("AGENTE", "ID"), ("COD", "AGENTE")], fallback=0)
    agent_name_col = find_col(headers_norm, [("NOMBRE", "AGENTE"), ("AGENTE",)], fallback=1)
    client_id_col = find_col(headers_norm, [("CLIENTE", "ID"), ("COD", "CLIENTE")], fallback=2)
    client_name_col = find_col(headers_norm, [("NOMBRE", "CLIENTE"), ("RAZON", "SOCIAL")], fallback=3)
    month_col = find_col(headers_norm, [("MES", "NUM"), ("NUM", "MES")], fallback=2)
    importe_col = find_col(
        headers_norm,
        [("IMPORTE", "NETO"), ("IMPORTE", "TOTAL"), ("IMPORTE",), ("VENTA",), ("FACTURACION",)],
        fallback=9,
    )
    beneficio_col = find_col(headers_norm, [("BENEFICIO",), ("GANANCIA",), ("PROFIT",)], fallback=5)

    agents_sales = {}
    client_rows = 0
    max_col = len(usable_headers)

    month_key_map = {
        2: "FEB",
        3: "MAR",
        4: "ABR",
        5: "MAY",
        6: "JUN",
        7: "JUL",
        8: "AGO",
        9: "SEP",
        10: "OCT",
        11: "NOV",
        12: "DIC",
        1: "ENE",
    }

    for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
        if not row:
            continue

        agent_id = as_int(row[agent_id_col] if agent_id_col < len(row) else None)
        if agent_id is None:
            continue

        client_rows += 1
        agent_name = as_str(row[agent_name_col] if agent_name_col < len(row) else f"Agente {agent_id}")
        client_id = as_int(row[client_id_col] if client_id_col < len(row) else None)
        client_name = as_str(row[client_name_col] if client_name_col < len(row) else "Cliente sin nombre")
        month_num = as_int(row[month_col] if month_col < len(row) else None)
        month_key = month_key_map.get(month_num, str(month_num))
        importe = as_float(row[importe_col] if importe_col < len(row) else None)
        beneficio = as_float(row[beneficio_col] if beneficio_col < len(row) else None)

        if agent_id not in agents_sales:
            agents_sales[agent_id] = {
                "name": agent_name,
                "monthly": {},
                "total_importe": 0.0,
                "total_beneficio": 0.0,
                "clients": set(),
            }

        agents_sales[agent_id]["total_importe"] += importe
        agents_sales[agent_id]["total_beneficio"] += beneficio
        agents_sales[agent_id]["monthly"][month_key] = agents_sales[agent_id]["monthly"].get(month_key, 0.0) + importe
        if client_id is not None:
            agents_sales[agent_id]["clients"].add(client_id)
        elif client_name:
            agents_sales[agent_id]["clients"].add(client_name)

    print(f"Total filas reales: {client_rows}")
    print(f"Agentes unicos en ventas reales: {len(agents_sales)}")
    for aid, info in sorted(agents_sales.items()):
        months = sorted(info["monthly"].keys())
        print(
            f"Agente {aid} ({info['name']}): "
            f"Venta Real 2026={info['total_importe']:.2f}, "
            f"Beneficio={info['total_beneficio']:.2f}, "
            f"Clientes={len(info['clients'])}, "
            f"Meses={months}"
        )


summarize_presupuestos("PRESUPUESTOS VENTAS.xlsx")
summarize_ventas_reales("VENTAS 2026.xlsx")
