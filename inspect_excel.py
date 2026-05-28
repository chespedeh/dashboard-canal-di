import openpyxl

def inspect_file(filename):
    print(f"=== Inspecting {filename} ===")
    wb = openpyxl.load_workbook(filename, read_only=True)
    print("Sheets:", wb.sheetnames)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"\nSheet: {sheetname}")
        # Get first 10 rows
        rows = list(sheet.iter_rows(max_row=10, values_only=True))
        for i, r in enumerate(rows):
            print(f"Row {i+1}: {r}")
    print("=============================\n")

inspect_file("PRESUPUESTOS VENTAS.xlsx")
inspect_file("VENTAS 2026.xlsx")
